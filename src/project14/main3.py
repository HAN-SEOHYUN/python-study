from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import dotenv
import os
dotenv_file = dotenv.find_dotenv()
dotenv.load_dotenv(dotenv_file)

load_wb = load_workbook(r"src/project14/이메일주소.xlsx", data_only=True)
load_ws = load_wb.active
# load_workbook : 엑셀파일을 로드하기 위해 사용하는 함수
# data_only=True : 수식이 아닌 셀의 값을 로드하도록 설정

for i in range(1, load_ws.max_row + 1):
    recv_email_value = load_ws.cell(i, 1).value
    print("성공:", recv_email_value)
    try:
        send_email = "h0031271@gmail.com"
        send_pwd = os.environ['GOOGLE_PASSWORD']

        recv_email = recv_email_value

        smtp_name = "smtp.gmail.com" # SMTP 서버 주소 할당
        smtp_port = 587 # SMTP 포트 번호 할당

        msg = MIMEMultipart()

        msg['Subject'] = "엑셀에서 메일 주소를 읽어 자동으로 보내는 메일입니다."
        msg['From'] = send_email
        msg['To'] = recv_email

        text = """
                메일내용입니다.
                감사합니다
                """
        msg.attach(MIMEText(text))

        s = smtplib.SMTP(smtp_name, smtp_port) # s : smtplib.SMTP 의 객체로, SMTP 서버와 연결을 담당함
        s.starttls() # SMTP 연결 암호화
        s.login(send_email, send_pwd) # SMTP 서버 로그인
        s.sendmail(send_email, recv_email, msg.as_string()) # 이메일을 발신자에서 수신자로 전송
        s.quit()
    except:
        print("에러:", recv_email_value)